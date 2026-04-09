#!/usr/bin/env python3
"""
JPX Market Alerts 자동 수집 스크립트
=============================================
1. Securities Under Supervision & Securities to Be Delisted
   - Current Designations (Stocks)        -> index.html
   - Designation History (Domestic Stocks) -> 01.html
   - Current Designations (Others)         -> 02.html
   - Designation History (Others)          -> 03.html

2. Issues in an Improvement Period, etc.
   - Companies in an Improvement Period          -> index.html (xlsx)
   - Companies subject to transitional measures  -> 01.html   (xlsx)

3. Issues in a Grace Period pertaining to Delisting Criteria, etc.
   - Current Designations (Stocks)        -> index.html
   - Designation History (Domestic Stocks) -> 01.html
   - Current Designations (Others)         -> 02.html  (may be empty)
   - Designation History (Others)          -> 03.html

출력 구조:
  data/
    supervision_delisting/
      current_stocks.csv / .json
      history_stocks.csv / .json   (전체 연도 archive 포함)
      current_others.csv / .json
      history_others.csv / .json   (전체 연도 archive 포함)
    improvement_period/
      companies_improvement_period.csv / .json
      companies_transitional_measures.csv / .json
    grace_period/
      current_stocks.csv / .json
      history_stocks.csv / .json   (전체 연도 archive 포함)
      current_others.csv / .json
      history_others.csv / .json   (전체 연도 archive 포함)
    pdfs/
      supervision_delisting/       <code>_<name>.pdf
      improvement_period/          <xlsx filename>.xlsx
      grace_period/                <code>_<name>.pdf, GracePeriod_EN.xlsx
"""

import json
import os
import re
import sys
from datetime import date

import fitz  # pymupdf
import requests
from bs4 import BeautifulSoup

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

BASE_URL = "https://www.jpx.co.jp"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
}
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(SCRIPT_DIR, "data")

# extract_pdf_reasons.py의 함수를 import
from extract_pdf_reasons import extract_reason, extract_text


def fetch_soup(path: str) -> BeautifulSoup:
    url = path if path.startswith("http") else BASE_URL + path
    r = requests.get(url, headers=HEADERS, timeout=30)
    r.raise_for_status()
    return BeautifulSoup(r.text, "html.parser")


def download_file(url: str, save_path: str) -> bool:
    if os.path.exists(save_path):
        print(f"  [SKIP] {os.path.basename(save_path)}")
        return True
    try:
        full_url = url if url.startswith("http") else BASE_URL + url
        r = requests.get(full_url, headers=HEADERS, timeout=60)
        r.raise_for_status()
        with open(save_path, "wb") as f:
            f.write(r.content)
        size_kb = len(r.content) / 1024
        print(f"  [OK] {os.path.basename(save_path)} ({size_kb:.0f} KB)")
        return True
    except Exception as e:
        print(f"  [FAIL] {os.path.basename(save_path)}: {e}")
        return False


def cleanup_dir(directory: str, keep_files: set[str]):
    """디렉토리 내에서 keep_files에 없는 파일 삭제 (current 기준 클린업)"""
    if not os.path.exists(directory):
        return
    for f in os.listdir(directory):
        fpath = os.path.join(directory, f)
        if os.path.isfile(fpath) and f not in keep_files:
            os.remove(fpath)
            print(f"  [DEL] {f} (no longer in current list)")


def download_and_cleanup(rows: list[dict], target_dir: str, is_current: bool = False):
    """rows에서 File_URL을 찾아 다운로드. is_current=True이면 목록에 없는 기존 파일 삭제."""
    downloaded_files = set()
    for row in rows:
        url = row.get("File_URL", "")
        if not url:
            continue
        ext = ".pdf" if ".pdf" in url else ".xlsx" if ".xlsx" in url else ""
        if not ext:
            continue
        code = row.get("Code", "unknown")
        name = sanitize_name(row.get("Issue Name", "unknown"))
        filename = f"{code}_{name}{ext}"
        downloaded_files.add(filename)
        download_file(url, os.path.join(target_dir, filename))

    if is_current:
        cleanup_dir(target_dir, downloaded_files)


def enrich_rows_with_pdf(rows: list[dict], pdf_dir: str, filename_fn=None):
    """각 행의 PDF를 다운로드하고, Details에 Reason 추출, File_Path에 상대경로 설정.
    filename_fn: row -> filename 변환 함수 (None이면 기본 {code}_{name}.pdf)
    """
    for row in rows:
        url = row.get("File_URL", "")
        if not url or ".pdf" not in url:
            continue

        if filename_fn:
            filename = filename_fn(row)
        else:
            code = row.get("Code", "unknown")
            name = sanitize_name(row.get("Issue Name", "unknown"))
            filename = f"{code}_{name}.pdf"

        filepath = os.path.join(pdf_dir, filename)
        download_file(url, filepath)

        # 상대경로 (프로젝트 루트 기준)
        row["File_Path"] = os.path.relpath(filepath, SCRIPT_DIR).replace("\\", "/")

        # Details가 비어있으면 PDF에서 Reason 추출
        if not row.get("Details", "").strip() and os.path.exists(filepath):
            try:
                text = extract_text(filepath)
                reason = extract_reason(text)
                if reason:
                    row["Details"] = reason
            except Exception:
                pass


def sanitize_name(name: str) -> str:
    """파일명에 사용 불가 문자 제거"""
    name = re.sub(r'[<>:"/\\|?*,]', "", name)
    name = re.sub(r"\s+", "_", name).strip("_")
    return name


def save_json(rows: list[dict], path: str):
    if not rows:
        return
    os.makedirs(os.path.dirname(path), exist_ok=True)
    output = {
        "fetched_date": date.today().isoformat(),
        "count": len(rows),
        "data": rows,
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f"  [JSON] {os.path.relpath(path, SCRIPT_DIR)} ({len(rows)} rows)")


def get_year_pages(soup) -> list[tuple[str, str]]:
    """dropdown에서 연도별 archive 페이지 목록 추출. [(year, path), ...]"""
    pages = []
    select = soup.find("select")
    if select:
        for opt in select.find_all("option"):
            year = opt.get_text(strip=True)
            path = opt.get("value", "")
            if path:
                pages.append((year, path))
    return pages


def scrape_history_all_years(base_path: str, with_sections: bool = False) -> list[dict]:
    """history 페이지의 모든 연도(dropdown archives)를 순회하여 데이터 수집"""
    soup = fetch_soup(base_path)
    year_pages = get_year_pages(soup)

    if not year_pages:
        # dropdown이 없으면 현재 페이지만
        year_pages = [("", base_path)]

    all_rows = []
    for year, path in year_pages:
        if path != base_path:
            soup = fetch_soup(path)
        tables = soup.find_all("table")
        for table in tables:
            if with_sections:
                prev_h3 = table.find_previous("h3")
                section = prev_h3.get_text(strip=True) if prev_h3 else ""
                rows = parse_table(table, section)
            else:
                rows = parse_table(table)
            if year:
                for row in rows:
                    row["Year"] = year
            all_rows.extend(rows)
        row_count = sum(len(t.find_all("tr")) - 1 for t in tables)
        if year:
            print(f"    {year}: {row_count} rows")

    return all_rows


def parse_table(table, section_name: str = "") -> list[dict]:
    """HTML 테이블을 dict 리스트로 파싱. PDF/xlsx 링크도 추출."""
    rows = table.find_all("tr")
    if not rows:
        return []

    # 헤더
    header_cells = rows[0].find_all("th")
    if not header_cells:
        header_cells = rows[0].find_all("td")
    col_names = [c.get_text(strip=True).replace("*", "").strip() for c in header_cells]

    data = []
    for tr in rows[1:]:
        cells = tr.find_all(["td", "th"])
        if len(cells) < len(col_names):
            continue
        row = {}
        if section_name:
            row["Section"] = section_name
        for i, col in enumerate(col_names):
            cell = cells[i] if i < len(cells) else None
            if cell:
                row[col] = cell.get_text(strip=True)
            else:
                row[col] = ""

        # PDF/xlsx 링크 추출
        for cell in cells:
            link = cell.find("a", href=lambda h: h and (".pdf" in str(h) or ".xlsx" in str(h)))
            if link:
                row["File_URL"] = link["href"]
                break

        data.append(row)
    return data


def xlsx_to_dicts(xlsx_path: str) -> list[dict]:
    """xlsx 파일을 dict 리스트로 변환.
    Improvement Period xlsx 특수 구조 처리:
    - 헤더는 'Code'가 포함된 행
    - 병합셀로 인한 빈 컬럼은 왼쪽 헤더의 'Deadline' 서브컬럼
    - 날짜값이 있으면 해당 기준 미달 + 개선기한을 의미
    """
    if not HAS_OPENPYXL:
        print(f"  [WARN] openpyxl not installed, cannot parse {xlsx_path}")
        return []
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    # 'Code'가 포함된 행을 헤더로 탐색
    header_row = None
    header_row_idx = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        vals = [str(c).strip() if c else "" for c in row]
        if "Code" in vals:
            header_row = vals
            header_row_idx = i
            break

    if not header_row:
        wb.close()
        return []

    # 병합셀로 인한 빈 컬럼 처리: 빈 컬럼은 스킵 (병합 잔여물)
    valid_cols = [(i, col) for i, col in enumerate(header_row)
                  if col and col != "None"]

    # 기준 컬럼 목록 (날짜가 들어오면 해당 기준 미달 + 개선기한을 의미)
    criteria_cols = {
        "Number of Shareholders", "Number of Tradable Shares",
        "Tradable Share Ratio", "Tradable Share Market Capitalization",
        "Trading Value", "Trading Volume", "Market Capitalization",
        "Amount of Net Assets",
    }

    # 데이터 행 읽기 (헤더 이후)
    data = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i <= header_row_idx:
            continue
        vals = [str(c).strip() if c is not None else "" for c in row]
        if not any(vals):
            continue

        d = {}
        violated = []  # 미달 기준 목록
        for idx, col in valid_cols:
            raw = vals[idx] if idx < len(vals) else ""
            # datetime 문자열 → YYYY-MM-DD
            if "00:00:00" in raw:
                raw = raw.split(" ")[0]
            # 기준 컬럼에 날짜가 있으면 → 미달 기준 + 개선기한
            if col in criteria_cols and raw:
                violated.append(col)
                d[f"{col} Deadline"] = raw
            else:
                d[col] = raw

        # Code가 비어있으면 데이터 행이 아님
        if not d.get("Code", "").strip():
            continue

        # 미달 기준 요약 필드
        if violated:
            d["Violated Criteria"] = ", ".join(violated)

        data.append(d)

    wb.close()
    return data


# ============================================================
# 1. Securities Under Supervision & Securities to Be Delisted
# ============================================================
def scrape_supervision():
    print("\n" + "=" * 60)
    print("1. Securities Under Supervision & Securities to Be Delisted")
    print("=" * 60)

    out_dir = os.path.join(DATA_DIR, "supervision_delisting")
    pdf_dir = os.path.join(DATA_DIR, "pdfs", "supervision_delisting")
    os.makedirs(pdf_dir, exist_ok=True)

    base = "/english/listing/market-alerts/supervision"

    # 모든 탭의 파일을 추적 (마지막에 클린업)
    all_known_files = set()

    def _track(rows):
        for row in rows:
            fp = row.get("File_Path", "")
            if fp:
                all_known_files.add(os.path.basename(fp))

    # --- Current Stocks (index.html) ---
    print("\n--- Current Designations (Stocks) ---")
    soup = fetch_soup(f"{base}/index.html")
    tables = soup.find_all("table")
    all_rows = []
    for table in tables:
        prev_h3 = table.find_previous("h3")
        section = prev_h3.get_text(strip=True) if prev_h3 else ""
        rows = parse_table(table, section)
        all_rows.extend(rows)

    enrich_rows_with_pdf(all_rows, pdf_dir)
    _track(all_rows)
    save_json(all_rows, os.path.join(out_dir, "current_stocks.json"))

    # --- History Stocks (01.html + archives) ---
    print("\n--- Designation History (Domestic Stocks) [all years] ---")
    all_rows = scrape_history_all_years(f"{base}/01.html")

    enrich_rows_with_pdf(all_rows, pdf_dir)
    _track(all_rows)
    save_json(all_rows, os.path.join(out_dir, "history_stocks.json"))

    # --- Current Others (02.html) ---
    print("\n--- Current Designations (Others) ---")
    soup = fetch_soup(f"{base}/02.html")
    tables = soup.find_all("table")
    all_rows = []
    for table in tables:
        prev_h3 = table.find_previous("h3")
        section = prev_h3.get_text(strip=True) if prev_h3 else ""
        rows = parse_table(table, section)
        all_rows.extend(rows)

    enrich_rows_with_pdf(all_rows, pdf_dir)
    _track(all_rows)
    save_json(all_rows, os.path.join(out_dir, "current_others.json"))

    # --- History Others (03.html + archives) ---
    print("\n--- Designation History (Others) [all years] ---")
    all_rows = scrape_history_all_years(f"{base}/03.html")

    enrich_rows_with_pdf(all_rows, pdf_dir)
    _track(all_rows)
    save_json(all_rows, os.path.join(out_dir, "history_others.json"))

    # 모든 탭 처리 후 클린업
    cleanup_dir(pdf_dir, all_known_files)


# ============================================================
# 2. Issues in an Improvement Period
# ============================================================
def scrape_improvement_period():
    print("\n" + "=" * 60)
    print("2. Issues in an Improvement Period")
    print("=" * 60)

    out_dir = os.path.join(DATA_DIR, "improvement_period")
    file_dir = os.path.join(DATA_DIR, "pdfs", "improvement_period")
    os.makedirs(file_dir, exist_ok=True)

    base = "/english/listing/market-alerts/improvement-period"

    # current xlsx 파일 추적 (클린업용)
    current_xlsx_files = set()

    # --- Companies in an Improvement Period (index.html) ---
    print("\n--- Companies in an Improvement Period ---")
    soup = fetch_soup(f"{base}/index.html")
    xlsx_links = soup.find_all("a", href=lambda h: h and ".xlsx" in str(h))
    for link in xlsx_links:
        href = link["href"]
        filename = os.path.basename(href)
        current_xlsx_files.add(filename)
        xlsx_path = os.path.join(file_dir, filename)
        download_file(href, xlsx_path)
        data = xlsx_to_dicts(xlsx_path)
        if data:
            save_json(data, os.path.join(out_dir, "companies_improvement_period.json"))

    # --- Companies subject to transitional measures (01.html) ---
    print("\n--- Companies subject to transitional measures ---")
    soup = fetch_soup(f"{base}/01.html")
    xlsx_links = soup.find_all("a", href=lambda h: h and ".xlsx" in str(h))
    for link in xlsx_links:
        href = link["href"]
        filename = os.path.basename(href)
        current_xlsx_files.add(filename)
        xlsx_path = os.path.join(file_dir, filename)
        download_file(href, xlsx_path)
        data = xlsx_to_dicts(xlsx_path)
        if data:
            save_json(data, os.path.join(out_dir, "companies_transitional_measures.json"))

    # 오래된 xlsx 클린업
    cleanup_dir(file_dir, current_xlsx_files)


# ============================================================
# 3. Issues in a Grace Period
# ============================================================
def scrape_grace_period():
    print("\n" + "=" * 60)
    print("3. Issues in a Grace Period")
    print("=" * 60)

    out_dir = os.path.join(DATA_DIR, "grace_period")
    pdf_dir = os.path.join(DATA_DIR, "pdfs", "grace_period")
    os.makedirs(pdf_dir, exist_ok=True)

    base = "/english/listing/market-alerts/grace-period"

    # current 파일 추적 (클린업용)
    current_files = set()

    # --- Current Stocks (index.html) ---
    print("\n--- Current Designations (Stocks) ---")
    soup = fetch_soup(f"{base}/index.html")
    tables = soup.find_all("table")
    all_rows = []
    for table in tables:
        prev_h3 = table.find_previous("h3")
        section = prev_h3.get_text(strip=True) if prev_h3 else ""

        # xlsx 링크 처리
        xlsx_link = table.find("a", href=lambda h: h and ".xlsx" in str(h))
        if xlsx_link:
            href = xlsx_link["href"]
            filename = os.path.basename(href)
            current_files.add(filename)
            xlsx_path = os.path.join(pdf_dir, filename)
            download_file(href, xlsx_path)
            xlsx_data = xlsx_to_dicts(xlsx_path)
            if xlsx_data:
                for row in xlsx_data:
                    row["Section"] = section
                all_rows.extend(xlsx_data)
            continue

        rows = parse_table(table, section)
        all_rows.extend(rows)

    enrich_rows_with_pdf(all_rows, pdf_dir)
    for row in all_rows:
        fp = row.get("File_Path", "")
        if fp:
            current_files.add(os.path.basename(fp))

    save_json(all_rows, os.path.join(out_dir, "current_stocks.json"))

    # --- History Stocks (01.html + archives) ---
    print("\n--- Designation History (Domestic Stocks) [all years] ---")
    all_rows = scrape_history_all_years(f"{base}/01.html")

    enrich_rows_with_pdf(all_rows, pdf_dir)
    for row in all_rows:
        fp = row.get("File_Path", "")
        if fp:
            current_files.add(os.path.basename(fp))

    save_json(all_rows, os.path.join(out_dir, "history_stocks.json"))

    # --- Current Others (02.html) ---
    print("\n--- Current Designations (Others) ---")
    soup = fetch_soup(f"{base}/02.html")
    tables = soup.find_all("table")
    all_rows = []
    for table in tables:
        prev_h3 = table.find_previous("h3")
        section = prev_h3.get_text(strip=True) if prev_h3 else ""
        rows = parse_table(table, section)
        all_rows.extend(rows)

    enrich_rows_with_pdf(all_rows, pdf_dir)
    for row in all_rows:
        fp = row.get("File_Path", "")
        if fp:
            current_files.add(os.path.basename(fp))

    if all_rows:
        save_json(all_rows, os.path.join(out_dir, "current_others.json"))
    else:
        print("  [INFO] No data (page may be empty)")

    # --- History Others (03.html + archives) ---
    print("\n--- Designation History (Others) [all years] ---")
    all_rows = scrape_history_all_years(f"{base}/03.html")

    enrich_rows_with_pdf(all_rows, pdf_dir)
    for row in all_rows:
        fp = row.get("File_Path", "")
        if fp:
            current_files.add(os.path.basename(fp))

    save_json(all_rows, os.path.join(out_dir, "history_others.json"))

    # 모든 탭 처리 후 클린업
    cleanup_dir(pdf_dir, current_files)


def _year_filename(row):
    """year_{code}_{name}.pdf 파일명 생성"""
    code = row.get("Code", "unknown")
    name = sanitize_name(row.get("Issue Name", "unknown"))
    year = row.get("Year", "")
    return f"{year}_{code}_{name}.pdf" if year else f"{code}_{name}.pdf"


# ============================================================
# 4. Public Announcement Measures
# ============================================================
def scrape_public_announcement():
    print("\n" + "=" * 60)
    print("4. Public Announcement Measures")
    print("=" * 60)

    out_dir = os.path.join(DATA_DIR, "public_announcement")
    pdf_dir = os.path.join(DATA_DIR, "pdfs", "public_announcement")
    os.makedirs(pdf_dir, exist_ok=True)

    base = "/english/listing/measures/public-announce"

    print("\n--- Public Announcement [all years] ---")
    all_rows = scrape_history_all_years(f"{base}/index.html")

    enrich_rows_with_pdf(all_rows, pdf_dir, filename_fn=_year_filename)
    all_known_files = {os.path.basename(r["File_Path"]) for r in all_rows if r.get("File_Path")}

    save_json(all_rows, os.path.join(out_dir, "public_announcement.json"))
    cleanup_dir(pdf_dir, all_known_files)


# ============================================================
# 5. Listing Agreement Violation Penalties
# ============================================================
def scrape_violation_penalties():
    print("\n" + "=" * 60)
    print("5. Listing Agreement Violation Penalties")
    print("=" * 60)

    out_dir = os.path.join(DATA_DIR, "violation_penalties")
    pdf_dir = os.path.join(DATA_DIR, "pdfs", "violation_penalties")
    os.makedirs(pdf_dir, exist_ok=True)

    base = "/english/listing/measures/listing-agreement-violation"

    print("\n--- Violation Penalties [all years] ---")
    all_rows = scrape_history_all_years(f"{base}/index.html")

    enrich_rows_with_pdf(all_rows, pdf_dir, filename_fn=_year_filename)
    all_known_files = {os.path.basename(r["File_Path"]) for r in all_rows if r.get("File_Path")}

    save_json(all_rows, os.path.join(out_dir, "violation_penalties.json"))
    cleanup_dir(pdf_dir, all_known_files)


# ============================================================
# 6. Companies Requested to Submit Improvement Reports
# ============================================================
def scrape_improvement_reports():
    print("\n" + "=" * 60)
    print("6. Improvement Reports / Improvement Status Reports")
    print("=" * 60)

    out_dir = os.path.join(DATA_DIR, "improvement_reports")
    pdf_dir = os.path.join(DATA_DIR, "pdfs", "improvement_reports")
    os.makedirs(pdf_dir, exist_ok=True)

    base = "/english/listing/measures/improvement-reports"

    print("\n--- Improvement Reports ---")
    soup = fetch_soup(f"{base}/index.html")
    tables = soup.find_all("table")
    all_rows = []
    for table in tables:
        rows = parse_table(table)
        all_rows.extend(rows)

    # 이 페이지는 Submission Date 컬럼에 여러 PDF 링크가 있을 수 있음
    # parse_table은 첫 번째 링크만 잡으므로, 모든 PDF를 별도 추출
    all_known_files = set()
    for table in tables:
        for tr in table.find_all("tr")[1:]:
            cells = tr.find_all(["td", "th"])
            if len(cells) < 2:
                continue
            code = cells[1].get_text(strip=True) if len(cells) > 1 else "unknown"
            name_raw = cells[0].get_text(strip=True) if cells else "unknown"
            name = sanitize_name(name_raw)
            pdf_paths = []
            for link in tr.find_all("a", href=lambda h: h and ".pdf" in str(h)):
                url = link["href"]
                link_text = sanitize_name(link.get_text(strip=True))
                filename = f"{code}_{name}_{link_text}.pdf"
                all_known_files.add(filename)
                filepath = os.path.join(pdf_dir, filename)
                download_file(url, filepath)
                pdf_paths.append(os.path.relpath(filepath, SCRIPT_DIR).replace("\\", "/"))

            # 해당 code의 행을 찾아서 File_Path와 Details 설정
            for row in all_rows:
                if row.get("Code") == code:
                    if pdf_paths:
                        row["File_Path"] = pdf_paths[0]
                    # 첫 번째 PDF에서 Details 추출
                    if not row.get("Details", "").strip() and pdf_paths:
                        first_pdf = os.path.join(SCRIPT_DIR, pdf_paths[0])
                        if os.path.exists(first_pdf):
                            try:
                                text = extract_text(first_pdf)
                                reason = extract_reason(text)
                                if reason:
                                    row["Details"] = reason
                            except Exception:
                                pass
                    break

    save_json(all_rows, os.path.join(out_dir, "improvement_reports.json"))
    cleanup_dir(pdf_dir, all_known_files)


# ============================================================
# 7. Securities on Special Alert
# ============================================================
def scrape_special_alert():
    print("\n" + "=" * 60)
    print("7. Securities on Special Alert")
    print("=" * 60)

    out_dir = os.path.join(DATA_DIR, "special_alert")
    pdf_dir = os.path.join(DATA_DIR, "pdfs", "special_alert")
    os.makedirs(pdf_dir, exist_ok=True)

    base = "/english/listing/measures/alert"
    all_known_files = set()

    # --- Current Designations (index.html) ---
    print("\n--- Current Designations ---")
    soup = fetch_soup(f"{base}/index.html")
    tables = soup.find_all("table")
    all_rows = []
    for table in tables:
        rows = parse_table(table)
        all_rows.extend(rows)

    enrich_rows_with_pdf(all_rows, pdf_dir)
    for row in all_rows:
        fp = row.get("File_Path", "")
        if fp:
            all_known_files.add(os.path.basename(fp))

    save_json(all_rows, os.path.join(out_dir, "current.json"))

    # --- Designation History (01.html + archives) ---
    print("\n--- Designation History [all years] ---")
    all_rows = scrape_history_all_years(f"{base}/01.html")

    enrich_rows_with_pdf(all_rows, pdf_dir)
    for row in all_rows:
        fp = row.get("File_Path", "")
        if fp:
            all_known_files.add(os.path.basename(fp))

    save_json(all_rows, os.path.join(out_dir, "history.json"))
    cleanup_dir(pdf_dir, all_known_files)


def main():
    print("JPX Market Alerts Scraper")
    print(f"Output: {DATA_DIR}")

    if not HAS_OPENPYXL:
        print("\n[WARN] openpyxl not installed. XLSX files will be downloaded but not parsed.")
        print("       Install with: pip install openpyxl\n")

    scrape_supervision()
    scrape_improvement_period()
    scrape_grace_period()
    scrape_public_announcement()
    scrape_violation_penalties()
    scrape_improvement_reports()
    scrape_special_alert()

    print("\n" + "=" * 60)
    print("DONE")
    print("=" * 60)


if __name__ == "__main__":
    main()
